import tkinter as tk
from tkinter import messagebox
from tkinter import *
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

excelfile = "ScoreTracker.xlsx"

if not os.path.exists(excelfile):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    sheet.append(["Name", "Score", "Remarks"])
    workbook.save(excelfile)

def add_to_excel(name_insert, score_insert):
    try:
        workbook = load_workbook(excelfile)
        sheet = workbook["Sheet"]

        # Find and remove existing "Average" row and the blank row above it if present
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == "Average":
                sheet.delete_rows(row)  # remove Average row
                if sheet.cell(row=row - 1, column=1).value is None and \
                   sheet.cell(row=row - 1, column=2).value is None and \
                   sheet.cell(row=row - 1, column=3).value is None:
                    sheet.delete_rows(row - 1)  # remove blank row above Average if any
                break

        next_row = sheet.max_row + 1
        remarks = "Passed" if score_insert >= 75 else "Failed"

        sheet.cell(row=next_row, column=1).value = name_insert
        sheet.cell(row=next_row, column=2).value = score_insert
        sheet.cell(row=next_row, column=3).value = remarks

        # Calculate average
        scores = []
        for row in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row, column=2).value
            if isinstance(val, (int, float)):
                scores.append(val)

        if scores:
            avg_score = sum(scores) / len(scores)
            avg_remarks = "Passed" if avg_score >= 75 else "Failed"

            insert_row = sheet.max_row + 1
            sheet.insert_rows(insert_row)  # Insert 1 blank row
            avg_row = insert_row + 1

            sheet.cell(row=avg_row, column=1).value = "Average"
            sheet.cell(row=avg_row, column=2).value = round(avg_score, 2)
            sheet.cell(row=avg_row, column=3).value = avg_remarks

        workbook.save(excelfile)
        excel_format()

    except PermissionError:
        messagebox.showerror("File Error", f"Permission denied. Please close the Excel file if it is open.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def excel_format():
    wb = load_workbook("ScoreTracker.xlsx")
    ws = wb["Sheet"]

    for cell in ws[1]:
        cell.font(bold = True)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.velue else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save("scoreTracker.xlsx")

window = tk.Tk()
window.title("Score Tracker")
window.geometry("300x300")
window.configure(bg="light blue")

label = tk.Label(master=window, text="Score Tracker System", font=("Times New Roman", 16), bg="light blue")
label.pack(pady=27)

frame = tk.Frame(master=window, bg="light blue")
frame.pack(padx=5, pady=5)

name_label = tk.Label(master=frame, text="Student Name:", bg="light blue")
name_label.grid(row=0, column=0, padx=5, sticky="w")
name_entry = tk.Entry(master=frame, width=30, bg="light blue")
name_entry.grid(row=0, column=1, padx=5, pady=7)

score_label = tk.Label(master=frame, text="Score:", bg="light blue")
score_label.grid(row=1, column=0, padx=5, sticky="w")
score_entry = tk.Entry(master=frame, width=30, bg="light blue")
score_entry.grid(row=1, column=1, padx=5, pady=7)

def get_name_score():
    name_insert = name_entry.get().strip()
    score_insert = score_entry.get().strip()

    if not name_insert or not score_insert:
        messagebox.showwarning("Missing Data", "Please fill all the fields!")
        return
    try:
        score_num = int(score_insert)
    except ValueError:
        messagebox.showerror("Input Error", "Score must be a number")
        return

    add_to_excel(name_insert, score_num)
    messagebox.showinfo("Success", "Data added successfully!")

    name_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)

btn = tk.Button(master=window, text="Add", width=15, command=get_name_score, bg="sky blue")
btn.pack(pady=7)

window.mainloop()
