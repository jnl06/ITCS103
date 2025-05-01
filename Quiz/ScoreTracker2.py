from openpyxl import load_workbook
import tkinter as tk
from tkinter import *
from tkinter import messagebox


def save_to_excel():

    name_insert = name_entry.get().strip()
    score_insert = score_entry.get().strip()

    if not name_insert or not score_insert:
        messagebox.showwarning("Missing Data", "Please fill all the fields!")
        return
    try:
        score_num = int(score_insert)
    except ValueError:
        messagebox.showerror("Input Error", "Score must be a number")
        return

    save_to_excel(name_insert, score_num)
    messagebox.showinfo("Success", "Data added successfully!")

    wb = load_workbook("ScoreTracker.xlsx")
    ws = wb["Sheet"]
    next_row = ws.max_row + 1

    remarks = "Passed" if score_insert >= 75 else "Failed"

    ws.cell(row=next_row, column=1).value = name_insert
    ws.cell(row=next_row, column=2).value = score_insert
    ws.cell(row=next_row, column=3).value = remarks


    ws.append([name_insert, score_insert])
    wb.save("ScoreTracker.xlsx")

    messagebox.showinfo("Success", "Data saved successfully!")

    name_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)


window = tk.Tk()
window.title("Score Tracker")
window.geometry("300x300")
window.configure(bg="light blue")

label = tk.Label(master=window, text="Score Tracker System", font=("Times New Roman", 16), bg="light blue")
label.pack(pady=27)

frame = tk.Frame(master=window, bg="light blue")
frame.pack(padx=5, pady=5)

name_label = tk.Label(master=frame, text="Student Name:", bg="light blue")
name_label.grid(row=0, column=0, padx=5, sticky="w")
name_entry = tk.Entry(master=frame, width=30, bg="light blue")
name_entry.grid(row=0, column=1, padx=5, pady=7)

score_label = tk.Label(master=frame, text="Score:", bg="light blue")
score_label.grid(row=1, column=0, padx=5, sticky="w")
score_entry = tk.Entry(master=frame, width=30, bg="light blue")
score_entry.grid(row=1, column=1, padx=5, pady=7)

btn = tk.Button(master=window, text="Add", width=15, command=save_to_excel, bg="sky blue")
btn.pack(pady=7)


window.mainloop()